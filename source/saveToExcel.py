# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""
from covid import Covid
import matplotlib.pyplot as pyplot
import time
import pandas as pd
from openpyxl import load_workbook

count = 1;

localTime = time.asctime(time.localtime(time.time()))
covid = Covid()
worldCovid = Covid(source="worldometers")
countries = worldCovid.list_countries()
for i in countries:
    print("%3s" %str(count) + ' ' + i)
    count = count + 1
    
    
country = input("\nEnter your country name: ")

data = covid.get_status_by_country_name(country)

cadr = {
        key: data[key]
        for key in data.keys() & {"confirmed", "active", "deaths", "recovered"}
        }

n = list(cadr.keys())
v = list(cadr.values())
print(cadr)


# Create DataTable & write to excel file

df = pd.DataFrame({'Time' : [localTime],
                   'Confirmed' : [cadr.get('confirmed')],
                   'Active' : [cadr.get('active')],
                   'Deaths' : [cadr.get('deaths')],
                   'Recovered' : [cadr.get('recovered')]})





writer = pd.ExcelWriter('output.xlsx', engine='openpyxl')
writer.book = load_workbook('output.xlsx')
writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

reader = pd.read_excel(r'output.xlsx')

if country.lower() in writer.sheets:
    df.to_excel(writer,index=False,header=False,startrow=writer.sheets[country.lower()].max_row, sheet_name=country.lower())
else:
    df.to_excel(writer,index=False,header=True,startrow=0, sheet_name=country.lower())


writer.close()

# chart
pyplot.title(country)
pyplot.bar(range(len(cadr)), v, tick_label=n)
pyplot.show()
  